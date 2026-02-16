import os
from typing import Dict, List, Any, Optional, Union, Tuple

class ExcelHandler:
    """
    Excel文件处理类，支持Excel文件的读写和单元格更新操作
    
    Note:
        依赖openpyxl库
    """
    
    @staticmethod
    def _ensure_openpyxl() -> None:
        """
        确保openpyxl库已安装
        
        Raises:
            ImportError: openpyxl库未安装
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("请安装openpyxl库: pip install openpyxl")
    
    @staticmethod
    def read_excel(file_path: str, sheet_name: Optional[str] = None, 
                   skip_header: bool = False) -> List[List[Any]]:
        """
        读取Excel文件
        
        Args:
            file_path (str): Excel文件路径
            sheet_name (str): 工作表名称，默认为第一个工作表
            skip_header (bool): 是否跳过表头，默认为False
            
        Returns:
            List[List[Any]]: 二维列表，每行数据作为一个子列表
            
        Raises:
            FileNotFoundError: 文件不存在
            Exception: 读取文件失败
        """
        ExcelHandler._ensure_openpyxl()
        import openpyxl
        
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        data = []
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            if sheet_name:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active
            
            start_row = 2 if skip_header else 1
            for row in sheet.iter_rows(min_row=start_row, values_only=True):
                # 过滤空行
                if any(cell is not None for cell in row):
                    data.append(list(row))
            
            workbook.close()
            return data
        except Exception as e:
            raise Exception(f"读取Excel文件失败: {str(e)}")
    
    @staticmethod
    def write_excel(file_path: str, data: List[List[Any]], sheet_name: str = 'Sheet1', 
                    header: Optional[List[str]] = None) -> bool:
        """
        写入Excel文件
        
        Args:
            file_path (str): Excel文件路径
            data (List[List[Any]]): 要写入的数据，二维列表
            sheet_name (str): 工作表名称，默认为'Sheet1'
            header (List[str]): 表头，可选
            
        Returns:
            bool: 写入是否成功
            
        Raises:
            Exception: 写入文件失败
        """
        ExcelHandler._ensure_openpyxl()
        import openpyxl
        
        try:
            # 确保目录存在
            os.makedirs(os.path.dirname(file_path) if os.path.dirname(file_path) else '.', exist_ok=True)
            
            workbook = openpyxl.Workbook()
            
            # 删除默认的Sheet
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
            
            # 创建新工作表
            sheet = workbook.create_sheet(sheet_name)
            
            # 写入表头
            if header:
                sheet.append(header)
            
            # 写入数据
            for row in data:
                sheet.append(row)
            
            workbook.save(file_path)
            workbook.close()
            return True
        except Exception as e:
            raise Exception(f"写入Excel文件失败: {str(e)}")
    
    @staticmethod
    def read_excel_dict(file_path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        以字典形式读取Excel文件（使用表头作为键）
        
        Args:
            file_path (str): Excel文件路径
            sheet_name (str): 工作表名称，默认为第一个工作表
            
        Returns:
            List[Dict[str, Any]]: 字典列表，每个字典表示一行数据
            
        Raises:
            FileNotFoundError: 文件不存在
            Exception: 读取文件失败
        """
        ExcelHandler._ensure_openpyxl()
        import openpyxl
        
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        data = []
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            if sheet_name:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active
            
            # 获取表头
            header = []
            for cell in sheet[1]:
                if cell.value is not None:
                    header.append(str(cell.value))
                else:
                    header.append('')
            
            # 读取数据
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = {}
                for i, cell_value in enumerate(row):
                    if i < len(header):
                        row_data[header[i]] = cell_value
                # 过滤空行
                if any(value is not None for value in row_data.values()):
                    data.append(row_data)
            
            workbook.close()
            return data
        except Exception as e:
            raise Exception(f"读取Excel文件失败: {str(e)}")
    
    @staticmethod
    def write_excel_dict(file_path: str, data: List[Dict[str, Any]], sheet_name: str = 'Sheet1') -> bool:
        """
        以字典形式写入Excel文件
        
        Args:
            file_path (str): Excel文件路径
            data (List[Dict[str, Any]]): 要写入的数据，字典列表
            sheet_name (str): 工作表名称，默认为'Sheet1'
            
        Returns:
            bool: 写入是否成功
            
        Raises:
            Exception: 写入文件失败
        """
        if not data:
            return True
        
        ExcelHandler._ensure_openpyxl()
        import openpyxl
        
        try:
            # 确保目录存在
            os.makedirs(os.path.dirname(file_path) if os.path.dirname(file_path) else '.', exist_ok=True)
            
            workbook = openpyxl.Workbook()
            
            # 删除默认的Sheet
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
            
            # 创建新工作表
            sheet = workbook.create_sheet(sheet_name)
            
            # 获取所有键作为表头
            fieldnames = list(data[0].keys())
            sheet.append(fieldnames)
            
            # 写入数据
            for row in data:
                row_values = [row.get(key) for key in fieldnames]
                sheet.append(row_values)
            
            workbook.save(file_path)
            workbook.close()
            return True
        except Exception as e:
            raise Exception(f"写入Excel文件失败: {str(e)}")
    
    @staticmethod
    def update_excel_cell(file_path: str, sheet_name: Optional[str] = None, 
                         cell: str = 'A1', value: Any = None) -> bool:
        """
        更新Excel文件中的单个单元格
        
        Args:
            file_path (str): Excel文件路径
            sheet_name (str): 工作表名称，默认为第一个工作表
            cell (str): 单元格地址，如'A1'
            value (Any): 要设置的值
            
        Returns:
            bool: 更新是否成功
            
        Raises:
            FileNotFoundError: 文件不存在
            Exception: 更新失败
        """
        ExcelHandler._ensure_openpyxl()
        import openpyxl
        
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        try:
            workbook = openpyxl.load_workbook(file_path)
            if sheet_name:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active
            
            sheet[cell] = value
            workbook.save(file_path)
            workbook.close()
            return True
        except Exception as e:
            raise Exception(f"更新Excel单元格失败: {str(e)}")
